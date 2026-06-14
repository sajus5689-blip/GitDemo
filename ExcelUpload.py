from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
import openpyxl


def update_excel_data(file_path, coluname, rowName, newValue):
    book = openpyxl.load_workbook(file_path)
    sheet = book.active
    Dict = {}

    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row = 1, column= i).value == coluname:
            Dict["col"] = i

    for i in range(1, sheet.max_row +1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row = i, column= j).value == rowName:
                Dict["row"] = i


    sheet.cell(row= Dict["row"], column=Dict["col"]).value = newValue
    book.save(file_path)


file_path = "C:\\Users\\USER\\Downloads\\download.xlsx"
fruitName = "Apple"
newValue = "999"
service_obj = Service("C:\\Users\\USER\\OneDrive\\Documents\\Selenium Python_Udemy\\chromedriver-win64\\chromedriver-win64\\chromedriver.exe")
driver = webdriver.Chrome(service = service_obj)
driver.implicitly_wait(5)

driver.get("https://rahulshettyacademy.com/upload-download-test/")
driver.find_element(By.XPATH, "//button[@id='downloadButton']").click()

#edit excel with updated value
update_excel_data(file_path, "price", fruitName, newValue)

driver.find_element(By.CSS_SELECTOR, "#fileinput").send_keys(file_path)
wait = WebDriverWait(driver, 10)
wait.until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, "div[role='alert'] div:nth-child(2)")))
print(driver.find_element(By.CSS_SELECTOR, "div[role='alert'] div:nth-child(2)").text)

priceColumn = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
actualPrice = driver.find_element(By.XPATH, "//div[text()='"+fruitName+"']/parent::div/parent::div/div[@id='cell-"+priceColumn+"-undefined']").text
assert actualPrice == newValue